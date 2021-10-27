from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl.styles import Font
import openpyxl
import xlsxwriter
import os
import sys
import datetime
from PyQt5 import QtCore, QtGui, QtWidgets


class Ui_MainWindow(object):
    def __init__(self):
        super().__init__()
        # -- WINDOW SHAPE --
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(258, 740)
        MainWindow.setGeometry(3020, 140, 258, 740)
        MainWindow.setWindowIcon(QtGui.QIcon("icon.png"))

        # -- WIDGET LAYOUT --
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.outer_frame = QtWidgets.QFrame(self.centralwidget)
        self.outer_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.outer_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.outer_frame.setObjectName("outer_frame")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.outer_frame)
        self.gridLayout_3.setObjectName("gridLayout_3")

        # -- MAIN FRAME --------------------------------------------------------------------------------------------
        self.task_frame = QtWidgets.QFrame(self.outer_frame)
        self.task_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.task_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.task_frame.setObjectName("task_frame")
        self.formLayout_2 = QtWidgets.QFormLayout(self.task_frame)
        self.formLayout_2.setObjectName("formLayout_2")
        self.label_high_priority = QtWidgets.QLabel(self.task_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_high_priority.setFont(font)
        self.label_high_priority.setObjectName("label_high_priority")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.SpanningRole, self.label_high_priority)
        self.label_first_due = QtWidgets.QLabel(self.task_frame)
        self.label_first_due.setObjectName("label_first_due")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_first_due)
        self.first_dateEdit = QtWidgets.QDateEdit(self.task_frame)
        self.first_dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2021, 10, 1), QtCore.QTime(0, 0, 0)))
        self.first_dateEdit.setCalendarPopup(True)
        self.first_dateEdit.setObjectName("first_dateEdit")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.first_dateEdit)
        self.first_task_textEdit = QtWidgets.QPlainTextEdit(self.task_frame)
        self.first_task_textEdit.setObjectName("first_task_textEdit")
        self.formLayout_2.setWidget(2, QtWidgets.QFormLayout.SpanningRole, self.first_task_textEdit)
        self.setFirstDate_button = QtWidgets.QPushButton(self.task_frame)
        self.setFirstDate_button.setObjectName("setFirstDate_button")
        self.formLayout_2.setWidget(5, QtWidgets.QFormLayout.LabelRole, self.setFirstDate_button)
        self.first_task_button = QtWidgets.QPushButton(self.task_frame)
        self.first_task_button.setObjectName("first_task_button")
        self.formLayout_2.setWidget(5, QtWidgets.QFormLayout.FieldRole, self.first_task_button)
        self.label_regular_priority = QtWidgets.QLabel(self.task_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_regular_priority.setFont(font)
        self.label_regular_priority.setObjectName("label_regular_priority")
        self.formLayout_2.setWidget(6, QtWidgets.QFormLayout.SpanningRole, self.label_regular_priority)
        self.label_second_due = QtWidgets.QLabel(self.task_frame)
        self.label_second_due.setObjectName("label_second_due")
        self.formLayout_2.setWidget(7, QtWidgets.QFormLayout.LabelRole, self.label_second_due)
        self.second_dateEdit = QtWidgets.QDateEdit(self.task_frame)
        self.second_dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2021, 10, 1), QtCore.QTime(0, 0, 0)))
        self.second_dateEdit.setCalendarPopup(True)
        self.second_dateEdit.setObjectName("second_dateEdit")
        self.formLayout_2.setWidget(7, QtWidgets.QFormLayout.FieldRole, self.second_dateEdit)
        self.second_task_textEdit = QtWidgets.QPlainTextEdit(self.task_frame)
        self.second_task_textEdit.setObjectName("second_task_textEdit")
        self.formLayout_2.setWidget(8, QtWidgets.QFormLayout.SpanningRole, self.second_task_textEdit)
        self.setSecondDate_button = QtWidgets.QPushButton(self.task_frame)
        self.setSecondDate_button.setObjectName("setSecondDate_button")
        self.formLayout_2.setWidget(9, QtWidgets.QFormLayout.LabelRole, self.setSecondDate_button)
        self.second_task_button = QtWidgets.QPushButton(self.task_frame)
        self.second_task_button.setObjectName("second_task_button")
        self.formLayout_2.setWidget(9, QtWidgets.QFormLayout.FieldRole, self.second_task_button)
        self.label_low_priority = QtWidgets.QLabel(self.task_frame)
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_low_priority.setFont(font)
        self.label_low_priority.setObjectName("label_low_priority")
        self.formLayout_2.setWidget(10, QtWidgets.QFormLayout.SpanningRole, self.label_low_priority)
        self.label_third_due = QtWidgets.QLabel(self.task_frame)
        self.label_third_due.setObjectName("label_third_due")
        self.formLayout_2.setWidget(11, QtWidgets.QFormLayout.LabelRole, self.label_third_due)
        self.third_dateEdit = QtWidgets.QDateEdit(self.task_frame)
        self.third_dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2021, 10, 1), QtCore.QTime(0, 0, 0)))
        self.third_dateEdit.setCalendarPopup(True)
        self.third_dateEdit.setObjectName("third_dateEdit")
        self.formLayout_2.setWidget(11, QtWidgets.QFormLayout.FieldRole, self.third_dateEdit)
        self.third_task_textEdit = QtWidgets.QPlainTextEdit(self.task_frame)
        self.third_task_textEdit.setObjectName("third_task_textEdit")
        self.formLayout_2.setWidget(12, QtWidgets.QFormLayout.SpanningRole, self.third_task_textEdit)
        self.setThirdDate_button = QtWidgets.QPushButton(self.task_frame)
        self.setThirdDate_button.setObjectName("setThirdDate_button")
        self.formLayout_2.setWidget(13, QtWidgets.QFormLayout.LabelRole, self.setThirdDate_button)
        self.third_task_button = QtWidgets.QPushButton(self.task_frame)
        self.third_task_button.setObjectName("third_task_button")
        self.formLayout_2.setWidget(13, QtWidgets.QFormLayout.FieldRole, self.third_task_button)
        spacerItem = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.formLayout_2.setItem(14, QtWidgets.QFormLayout.LabelRole, spacerItem)
        self.label_reminder = QtWidgets.QLabel(self.task_frame)
        self.label_reminder.setObjectName("label_reminder")
        self.formLayout_2.setWidget(16, QtWidgets.QFormLayout.SpanningRole, self.label_reminder)
        self.reminder_lineEdit = QtWidgets.QLineEdit(self.task_frame)
        self.reminder_lineEdit.setText("")
        self.reminder_lineEdit.setObjectName("reminder_lineEdit")
        self.formLayout_2.setWidget(17, QtWidgets.QFormLayout.SpanningRole, self.reminder_lineEdit)
        self.reminder_lineEdit_2 = QtWidgets.QLineEdit(self.task_frame)
        self.reminder_lineEdit_2.setObjectName("reminder_lineEdit_2")
        self.formLayout_2.setWidget(18, QtWidgets.QFormLayout.SpanningRole, self.reminder_lineEdit_2)
        self.undo_button = QtWidgets.QPushButton(self.task_frame)
        self.undo_button.setObjectName("undo_button")
        self.formLayout_2.setWidget(19, QtWidgets.QFormLayout.LabelRole, self.undo_button)
        self.save_button = QtWidgets.QPushButton(self.task_frame)
        self.save_button.setObjectName("save_button")
        self.formLayout_2.setWidget(19, QtWidgets.QFormLayout.FieldRole, self.save_button)
        self.gridLayout_3.addWidget(self.task_frame, 0, 0, 1, 1)

        # -- BUTTON FRAME -- -----------------------------------------------------------------------------------
        self.button_frame = QtWidgets.QFrame(self.outer_frame)
        self.button_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.button_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        self.button_frame.setObjectName("button_frame")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.button_frame)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_control_title = QtWidgets.QLabel(self.button_frame)
        self.label_control_title.setObjectName("label_control_title")
        self.gridLayout_2.addWidget(self.label_control_title, 0, 0, 1, 2)
        self.drill_button = QtWidgets.QPushButton(self.button_frame)
        self.drill_button.setObjectName("drill_button")
        self.gridLayout_2.addWidget(self.drill_button, 1, 0, 1, 1)
        self.switch_button = QtWidgets.QPushButton(self.button_frame)
        self.switch_button.setObjectName("switch_button")
        self.gridLayout_2.addWidget(self.switch_button, 1, 1, 1, 1)
        self.bubble_button = QtWidgets.QPushButton(self.button_frame)
        self.bubble_button.setObjectName("bubble_button")
        self.gridLayout_2.addWidget(self.bubble_button, 2, 0, 1, 1)
        self.showArchive_button = QtWidgets.QPushButton(self.button_frame)
        self.showArchive_button.setObjectName("showArchive_button")
        self.gridLayout_2.addWidget(self.showArchive_button, 2, 1, 1, 1)
        self.gridLayout_3.addWidget(self.button_frame, 1, 0, 1, 1)

        # -- OUTER FRAME CONTENTS --
        self.label_status = QtWidgets.QLabel(self.outer_frame)
        self.label_status.setObjectName("label_status")
        self.gridLayout_3.addWidget(self.label_status, 2, 0, 1, 1)
        self.gridLayout.addWidget(self.outer_frame, 0, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 305, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        # -- CLASS VARIABLES --
        self.db_name = "task_db.xlsx"
        self.headers = ["A", "B", "C", "D", "E"]
        self.dues = [self.first_dateEdit, self.second_dateEdit, self.third_dateEdit]

        # -- INIT METHOD CALLS --
        self.first_setup()
        self.populate_current_tasks()

        # -- BUTTON CALLBACKS --
        self.showArchive_button.clicked.connect(self.showArchive)
        self.save_button.clicked.connect(self.savePriority)
        self.switch_button.clicked.connect(self.switchTask)
        self.drill_button.clicked.connect(self.drillTask)
        self.bubble_button.clicked.connect(self.bubbleTask)
        self.first_task_button.clicked.connect(self.firstTaskFinished)
        self.second_task_button.clicked.connect(self.secondTaskFinished)
        self.third_task_button.clicked.connect(self.thirdTaskFinished)
        self.undo_button.clicked.connect(self.undo)
        self.setFirstDate_button.clicked.connect(self.setFirstDate)
        self.setSecondDate_button.clicked.connect(self.setSecondDate)
        self.setThirdDate_button.clicked.connect(self.setThirdDate)

    # --- INIT SETUP FUNCTIONS --- They run every time
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Task Priority"))
        self.label_high_priority.setText(_translate("MainWindow", "Task with highest priority:"))
        self.label_first_due.setText(_translate("MainWindow", "Due:"))
        self.first_task_textEdit.setPlaceholderText(_translate("MainWindow", "Highest Priority Task"))
        self.setFirstDate_button.setText(_translate("MainWindow", "Set Date"))
        self.first_task_button.setText(_translate("MainWindow", "Task Finished"))
        self.label_regular_priority.setText(_translate("MainWindow", "Task with regular priority:"))
        self.label_second_due.setText(_translate("MainWindow", "Due:"))
        self.second_task_textEdit.setPlaceholderText(_translate("MainWindow", "Regular Priority Task"))
        self.setSecondDate_button.setText(_translate("MainWindow", "Set Date"))
        self.second_task_button.setText(_translate("MainWindow", "Task Finished"))
        self.label_low_priority.setText(_translate("MainWindow", "Task with low priority:"))
        self.label_third_due.setText(_translate("MainWindow", "Due:"))
        self.third_task_textEdit.setPlaceholderText(_translate("MainWindow", "Lowest Priority Task"))
        self.setThirdDate_button.setText(_translate("MainWindow", "Set Date"))
        self.third_task_button.setText(_translate("MainWindow", "Task Finished"))
        self.label_reminder.setText(_translate("MainWindow", "Reminders / Notes"))
        self.reminder_lineEdit.setPlaceholderText(_translate("MainWindow", "Note"))
        self.reminder_lineEdit_2.setPlaceholderText(_translate("MainWindow", "Note"))
        self.undo_button.setText(_translate("MainWindow", "Undo"))
        self.save_button.setText(_translate("MainWindow", "Save"))
        self.label_control_title.setText(_translate("MainWindow", "Priority Control"))
        self.drill_button.setText(_translate("MainWindow", "Drill (↓)"))
        self.switch_button.setText(_translate("MainWindow", "Switch (↓↑)"))
        self.bubble_button.setText(_translate("MainWindow", "Bubble (↑)"))
        self.showArchive_button.setText(_translate("MainWindow", "Show Archive"))
        self.label_status.setText(_translate("MainWindow", "Status"))

    def first_setup(self) -> None:
        """Checks whether its the first time setup in the local system
        If it is, it creates the local db"""
        if not os.path.isfile(self.db_name):
            print("No local db found - Creating local db")
            # -- CREATE LOCAL DB --
            workbook = xlsxwriter.Workbook(self.db_name)
            workbook.close()
            # -- FORMAT LOCAL DB --
            wb = openpyxl.load_workbook(self.db_name)
            ws = wb.active
            # - COLUMN HEADERS -
            ws["A1"], ws["B1"] = "TASK ID", "TASK DESCRIPTION"
            ws["C1"], ws["D1"] = "DATE ISSUED", "DATE FINISHED"
            ws["E1"], ws["F1"] = "DUE", "NOTES"
            # - SET HEADERS TO BOLD -
            for element in [ws["A1"], ws["B1"], ws["C1"], ws["D1"], ws["E1"], ws["F1"]]:
                element.font = openpyxl.styles.Font(bold=True)
            # - SET HEADERS WIDTH -
            for header in self.headers:
                ws.column_dimensions[header].width = 25
            ws.column_dimensions["A"].width = 15
            # - INIT EXAMPLE TASKS -
            now = datetime.datetime.now()
            ws["A2"], ws["A3"], ws["A4"] = 1, 2, 3
            ws["B2"], ws["B3"], ws["B4"] = "Example task 1", "Example task 2", "Example task 3"
            ws["C2"], ws["C3"], ws["C4"] = now, now, now
            ws["E2"], ws["E3"], ws["E4"] = False, False, False
            # - FREEZE HEADERS ROW -
            ws.freeze_panes = "A2"
            wb.save(self.db_name)
            wb.close()

    def populate_current_tasks(self) -> None:
        """Populate the task text edits with information pulled from the db"""
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # -- POPULATE FROM DB --
        id_list = [cell.value for cell in list(ws["A"])[-3:]]
        task_list = [cell.value for cell in list(ws["B"])[-3:]]
        dues_list = [cell.value for cell in list(ws["E"])[-3:]]
        notes_list = [cell.value for cell in list(ws["F"])[-3:]]
        # - POPULATE LABELS -
        self.label_high_priority.setText(f"Task with highest priority: #ID({str(id_list[0])})")
        self.label_regular_priority.setText(f"Task with regular priority: #ID({str(id_list[1])})")
        self.label_low_priority.setText(f"Task with lowest priority: #ID({str(id_list[2])})")
        # - POPULATE TEXT EDITS -
        self.first_task_textEdit.setPlainText(str(task_list[0]))
        self.second_task_textEdit.setPlainText(str(task_list[1]))
        self.third_task_textEdit.setPlainText(str(task_list[2]))
        # - POPULATE DATE EDITS -
        today = datetime.date.today()
        for (db_due, date_edit) in zip(dues_list, self.dues):
            if db_due is False:
                date_edit.setDate(today)
            else:
                date_edit.setDate(db_due)
                date_edit.setEnabled(False)
        # - POPULATE NOTES -
        print(f"Populating from notes_list at: {notes_list}")
        self.reminder_lineEdit.setText(notes_list[-3])
        self.reminder_lineEdit_2.setText(notes_list[-2])
        wb.close()
        print(f"Populated in order: {id_list[0], id_list[1], id_list[2]}")

    # --- UI MANIPULATION FUNCTIONS ---
    def setStatus(self, status) -> None:
        self.label_status.setText(status)

    def defaultStatus(self) -> None:
        self.label_status.setText("Status - Ok")

    def defaultPriorityLabels(self) -> None:
        self.label_high_priority.setText("Task with highest priority:")
        self.label_regular_priority.setText("Task with regular priority:")
        self.label_low_priority.setText("Task with lowest priority:")

    # --- SUPPORT FUNCTIONS ---
    def getIdsFromUI(self) -> list:  # UI priority ordered list
        top_id = self.label_high_priority.text().split("(")[1][:-1]
        regular_id = self.label_regular_priority.text().split("(")[1][:-1]
        lowest_id = self.label_low_priority.text().split("(")[1][:-1]
        return [top_id, regular_id, lowest_id]

    def getTasksFromUI(self) -> list:
        top_task = self.first_task_textEdit.toPlainText()
        regular_task = self.second_task_textEdit.toPlainText()
        lowest_task = self.third_task_textEdit.toPlainText()
        return [top_task, regular_task, lowest_task]

    def getIdsFromDB(self) -> list:
        """Gets Task ID from database, return list of sorted IDs"""
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # - GET ONLY LAST 3 VALUES -
        ids = [int(cell.value) for cell in list(ws["A"])[-3:]]
        wb.close()
        # - SORT -
        ids.sort()
        return ids

    # --- BUTTON CLICKED FUNCTIONS ---
    def switchTask(self) -> None:
        tasks = self.getTasksFromUI()
        self.first_task_textEdit.setPlainText(tasks[-1])
        self.third_task_textEdit.setPlainText(tasks[0])

        ids = self.getIdsFromUI()
        self.defaultPriorityLabels()
        self.label_high_priority.setText(f"Task with highest priority: #ID({ids[-1]})")
        self.label_regular_priority.setText(f"Task with regular priority: #ID({ids[1]})")
        self.label_low_priority.setText(f"Task with lowest priority: #ID({ids[0]})")

    def drillTask(self) -> None:
        tasks = self.getTasksFromUI()
        self.first_task_textEdit.setPlainText(tasks[1])
        self.second_task_textEdit.setPlainText(tasks[0])

        ids = self.getIdsFromUI()
        self.defaultPriorityLabels()
        self.label_high_priority.setText(f"Task with highest priority: #ID({ids[1]})")
        self.label_regular_priority.setText(f"Task with regular priority: #ID({ids[0]})")
        self.label_low_priority.setText(f"Task with lowest priority: #ID({ids[-1]})")

    def bubbleTask(self) -> None:
        tasks = self.getTasksFromUI()
        self.third_task_textEdit.setPlainText(tasks[1])
        self.second_task_textEdit.setPlainText(tasks[-1])

        ids = self.getIdsFromUI()
        self.defaultPriorityLabels()
        self.label_high_priority.setText(f"Task with highest priority: #ID({ids[0]})")
        self.label_regular_priority.setText(f"Task with regular priority: #ID({ids[-1]})")
        self.label_low_priority.setText(f"Task with lowest priority: #ID({ids[1]})")

    def savePriority(self) -> None:
        """Get the current priorities in the db, then get the current priorities in the ui and switch them"""
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # -- GET CURRENT PRIORITIES IN DB --
        db_ids = [cell.coordinate for cell in list(ws["A"])[-3:]]
        db_tasks = [cell.coordinate for cell in list(ws["B"])[-3:]]
        # -- CURRENT PRIORITIES IN UI --
        ui_ids = self.getIdsFromUI()
        print(f" Current IDs in UI: {ui_ids}")
        ui_tasks = self.getTasksFromUI()
        # -- MODIFY PRIORITIES IN DB --
        for coordinate, ui_id in zip(db_ids, ui_ids):
            ws[str(coordinate)] = ui_id
        for coordinate, ui_task in zip(db_tasks, ui_tasks):
            ws[str(coordinate)] = ui_task
        # -- SAVE NOTES IN DB --
        ws["F" + str(ws.max_row - 1)] = self.reminder_lineEdit.text()
        ws["F" + str(ws.max_row)] = self.reminder_lineEdit_2.text()

        wb.save(self.db_name)
        wb.close()

    def showArchive(self) -> None:
        os.startfile(self.db_name)

    def firstTaskFinished(self) -> None:
        """Adds a new task at the end of db and repopulates tasks"""
        self.savePriority()
        # -- GET NEW TASK ID --
        highest_id = self.getIdsFromDB()[-1]
        # -- WRITE NEW ID AND TASK --
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # -- LOG EXITING TASK WITH TIMESTAMP
        ws["D" + str(ws.max_row - 2)] = datetime.datetime.now()
        # --PUSH NEW TASK IN AND POPULATE DB --
        ws["A" + str(ws.max_row + 1)] = int(highest_id) + 1
        ws["B" + str(ws.max_row)] = "New task"
        ws["C" + str(ws.max_row)] = datetime.datetime.now()
        ws["E" + str(ws.max_row)] = False
        # -- RESET DATE EDIT --
        self.first_dateEdit.setEnabled(True)
        wb.save(self.db_name)
        wb.close()
        self.populate_current_tasks()

    def secondTaskFinished(self) -> None:
        """If second task i.e. regular priority task is finished, we use the priority control functions to
        first drill down the highest priority and allow the regular priority to exit the priority queue"""
        self.savePriority()
        # -- GET NEW TASK ID --
        highest_id = max(self.getIdsFromDB())
        print(self.getIdsFromDB(), "Drilling")
        self.drillTask()
        self.savePriority()
        # -- WRITE NEW ID AND TASK --
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # -- LOG EXITING TASK WITH TIMESTAMP
        ws["D" + str(ws.max_row - 2)] = datetime.datetime.now()
        # --PUSH NEW TASK IN AND POPULATE DB --
        ws["A" + str(ws.max_row + 1)] = int(highest_id) + 1
        ws["B" + str(ws.max_row)] = "New task"
        ws["C" + str(ws.max_row)] = datetime.datetime.now()
        ws["E" + str(ws.max_row)] = False
        # -- RESET DATE EDIT --
        self.second_dateEdit.setEnabled(True)
        wb.save(self.db_name)
        wb.close()
        self.populate_current_tasks()

    def thirdTaskFinished(self) -> None:
        """If third task i.e. lowest priority task is finished, we use the priority control functions to
        first switch the highest priority and the lowest, and allow the lowest priority to exit the priority queue,
        Then we change them back"""
        self.savePriority()
        # -- GET NEW TASK ID --
        highest_id = max(self.getIdsFromDB())
        print(highest_id)
        print(self.getIdsFromDB(), "Switching")
        self.switchTask()
        self.savePriority()
        # -- WRITE NEW ID AND TASK --
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        # -- LOG EXITING TASK WITH TIMESTAMP
        ws["D" + str(ws.max_row - 2)] = datetime.datetime.now()
        # --PUSH NEW TASK IN AND POPULATE DB --
        ws["A" + str(ws.max_row + 1)] = int(highest_id) + 1
        ws["B" + str(ws.max_row)] = "New task"
        ws["C" + str(ws.max_row)] = datetime.datetime.now()
        ws["E" + str(ws.max_row)] = False
        # -- RESET DATE EDIT --
        self.third_dateEdit.setEnabled(True)
        wb.save(self.db_name)
        wb.close()
        self.populate_current_tasks()
        self.drillTask()
        self.savePriority()

    def undo(self) -> None:
        """Retrieve task ID and task from temp file and restates it to the UI"""
        # -- DELETE NEW TASK --
        wb = openpyxl.load_workbook(self.db_name)
        ws = wb.active
        if ws["B" + str(ws.max_row)].value == "New task":
            ws.delete_rows(ws.max_row, 1)
        wb.save(self.db_name)
        wb.close()
        self.populate_current_tasks()

    def setFirstDate(self) -> None:
        """Checks if date edit is enabled, if it is, then set its current assigned UI value to DB,
        If it isn't, them enable it so user can modify its UI value"""
        if self.first_dateEdit.isEnabled():
            wb = openpyxl.load_workbook(self.db_name)
            ws = wb.active
            # -- GET UI VALUE --
            date = self.first_dateEdit.date()
            # -- WRITE UI VALUE TO DB --
            ws["E" + str(ws.max_row - 2)] = date.toPyDate()
            wb.save(self.db_name)
            wb.close()
            self.first_dateEdit.setEnabled(False)
        else:
            self.first_dateEdit.setEnabled(True)

    def setSecondDate(self) -> None:
        """Checks if date edit is enabled, if it is, then set its current assigned UI value to DB,
        If it isn't, them enable it so user can modify its UI value"""
        if self.second_dateEdit.isEnabled():
            wb = openpyxl.load_workbook(self.db_name)
            ws = wb.active
            # -- GET UI VALUE --
            date = self.second_dateEdit.date()
            # -- WRITE UI VALUE TO DB --
            ws["E" + str(ws.max_row - 1)] = date.toPyDate()
            wb.save(self.db_name)
            wb.close()
            self.second_dateEdit.setEnabled(False)
        else:
            self.second_dateEdit.setEnabled(True)

    def setThirdDate(self) -> None:
        """Checks if date edit is enabled, if it is, then set its current assigned UI value to DB,
        If it isn't, them enable it so user can modify its UI value"""
        if self.third_dateEdit.isEnabled():
            wb = openpyxl.load_workbook(self.db_name)
            ws = wb.active
            # -- GET UI VALUE --
            date = self.third_dateEdit.date()
            # -- WRITE UI VALUE TO DB --
            ws["E" + str(ws.max_row)] = date.toPyDate()
            wb.save(self.db_name)
            wb.close()
            self.third_dateEdit.setEnabled(False)
        else:
            self.third_dateEdit.setEnabled(True)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    MainWindow.show()
    sys.exit(app.exec_())
